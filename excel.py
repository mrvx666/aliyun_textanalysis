# -*- coding: UTF-8 -*-
import xlrd,xlwt,xlutils
import openpyxl
from openpyxl import workbook as  wb


def open_workbook(filename):
    return openpyxl.load_workbook(filename=filename)


def read_rawdata_arr(file):
    data_arr = []
    sheet = read_rawdata(file)
    datatuple = list(sheet["A"])
    for cell in datatuple:
        data_arr.append(cell.value)
    return data_arr


def read_rawdata(file):
    print('open excel success!')
    wb1 = openpyxl.load_workbook(filename=file)
    #sheetlist = wb1.get_sheet_names()
    #sheet = wb1.get_sheet_by_name(sheetlist[0])
    sheet = wb1.active
    return sheet


def write_to_excel(file, row, col, data):
    wb1 = openpyxl.Workbook()
    sheet = wb1.active
    sheet.cell(row=row+1, column=col+1).value = data
    #print("write data to row:" + str(row) + " column:" + str(col) + " success")
    wb1.save(filename=file)
    return 0


#write_to_excel('newdata.xlsx', 2, 2, 12345678)

