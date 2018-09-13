# -*- coding: UTF-8 -*-
from aliws import aliws
from attention_text_summary import attention_text_summary
from key_sentences_extraction import key_sentences_extraction
from key_words_extraction import key_words_extraction
from news_element_extraction import news_element_extraction
from excel import read_rawdata_arr ,open_workbook,write_to_excel
from address_analysis import address_extraction_arr

import openpyxl


def write_aliws(ws, rawdataarr, limit, ex):
    row = 1
    column = 1
    print(rawdataarr)
    for i in range(0, limit, 1):
        data = rawdataarr[i]
        dataarr = aliws(data)
        print(dataarr)

        for j in range(0, len(dataarr), 1):
            datadict = dataarr[j]
            ws.cell(row=row, column=column).value = datadict['word']
            print('write values {} to {},{} success!'.format(datadict['word'], row, column))
            ws.cell(row=row+1, column=column).value = datadict['wtype']
            print('write values {} to {},{} success!'.format(datadict['wtype'], row+1, column))

            column += 1

        row += 2
        column = 1
        ex.save("newdata.xlsx")


def write_key_words_extraction(ws, rawdataarr, limit, ex):
    row = 1
    column = 1
    print(rawdataarr)
    for i in range(0, limit, 1):
        data = rawdataarr[i]
        dataarr = key_words_extraction(data)
        print(dataarr)
        ws.cell(row=row, column=column).value = data
        print('write data {} to {},{} success!'.format(data, row, column))
        for j in range(0, len(dataarr), 1):
            datadict = dataarr[j]
            ws.cell(row=row, column=column+1).value = datadict['word']
            print('write values {} to {},{} success!'.format(datadict['word'], row, column+1))
            ws.cell(row=row + 1, column=column+1).value = datadict['weight']
            print('write values {} to {},{} success!'.format(datadict['weight'], row + 1, column+1))

            column += 1

        row += 2
        column = 1
        ex.save("newdata.xlsx")


def write_key_sentences_extraction(ws, rawdataarr, limit, ex):
    row = 1
    column = 1
    print(rawdataarr)
    for i in range(0, limit, 1):
        data = rawdataarr[i]
        dataarr = key_sentences_extraction(data)
        print(dataarr)
        ws.cell(row=row, column=column).value = data
        print('write data {} to {},{} success!'.format(data, row, column))
        for j in range(0, len(dataarr), 1):
            datadict = dataarr[j]
            ws.cell(row=row, column=column+1).value = datadict['sentence']
            print('write values {} to {},{} success!'.format(datadict['sentence'], row, column))
            ws.cell(row=row + 1, column=column+1).value = datadict['weight']
            print('write values {} to {},{} success!'.format(datadict['weight'], row + 1, column))

            column += 1

        row += 2
        column = 1
        ex.save("newdata.xlsx")


def write_attention_text_summary(ws, rawdataarr, limit, ex):
    row = 1
    column = 1
    print(rawdataarr)
    for i in range(0, limit, 1):
        data = rawdataarr[i]
        summary = attention_text_summary(data)
        ws.cell(row=row, column=column).value = summary
        print('write values {} to {},{} success!'.format(summary, row, column))
        row += 1
        ex.save("newdata.xlsx")


def write_news_element_extraction(ws, rawdataarr, limit, ex):
    row = 1
    column = 1
    print(rawdataarr)
    for i in range(0, limit, 1):
        data = rawdataarr[i]
        ws.cell(row=row, column=column).value = data
        newsdict = news_element_extraction(data)
        print(newsdict)
        for key in newsdict:
            dataarr = newsdict[key]
            #print(dataarr)
            for j in range(0, len(dataarr), 1):
                datadict = dataarr[j]
                ws.cell(row=row, column=column+1).value = datadict['tag']
                print('write values {} to {},{} success!'.format(datadict['tag'], row, column+1))
                ws.cell(row=row + 1, column=column+1).value = datadict['weight']
                print('write values {} to {},{} success!'.format(datadict['weight'], row + 1, column+1))

                column += 1
                print("now row is "+str(row))
            row += 2
            column = 1

        column = 1
        ex.save("newdata.xlsx")


def main():
    rawdata = "weibo_weibo_datatxt.xlsx"
    rawtablearr = read_rawdata_arr(rawdata)
    LIMIT = len(rawtablearr)
    resultarr = address_extraction_arr(rawtablearr)
    print(resultarr)
    print(len(resultarr))
    i = 0
    j = 0
    wb1 = openpyxl.Workbook()
    sheet = wb1.active
    for data in resultarr:
        print(data)
        sheet.cell(i+1, j+1).value = data
        i = i+1
    wb1.save("newdata.xlsx")


def col_maker_10():

    rawtablearr = read_rawdata_arr()
    limit = len(rawtablearr)
    wb = openpyxl.load_workbook(filename='newdata.xlsx')
    print('open excel success!')
    sheetlist = wb.get_sheet_names()
    ws = ex.get_sheet_by_name(sheetlist[0])
    print('open sheet1 success!')
    #1
    row = 0
    col = 0
    for i in range(0, 8190, 1):

        ws.cell(row=row + 1, column=col + 1).value = i
        #print('write values {} to {},{} success!'.format(i, row, col))
        row += 1
    #2
    row = 0
    col = 0
    for i in range(0, 8190, 1):
        ws.cell(row=row + 1, column=col + 2).value = i+0.1
        row += 1
    #3
    row = 0
    col = 0
    for i in range(0, 8190, 1):
        ws.cell(row=row + 1, column=col + 3).value = i + 0.2
        row += 1
    #4
    row = 0
    col = 0
    for i in range(0, 8190, 1):
        ws.cell(row=row + 1, column=col + 4).value = i + 0.3
        row += 1
    #5
    row = 0
    col = 0
    for i in range(0, 8190, 1):
        ws.cell(row=row + 1, column=col + 5).value = i + 0.4
        row += 1
    #6
    row = 0
    col = 0
    for i in range(0, 8190, 1):
        ws.cell(row=row + 1, column=col + 6).value = i + 0.5
        row += 1
    #7
    row = 0
    col = 0
    for i in range(0, 8190, 1):
        ws.cell(row=row + 1, column=col + 7).value = i + 0.6
        row += 1
    #8
    row = 0
    col = 0
    for i in range(0, 8190, 1):
        ws.cell(row=row + 1, column=col + 8).value = i + 0.7
        row += 1
    #9
    row = 0
    col = 0
    for i in range(0, 8190, 1):
        ws.cell(row=row + 1, column=col + 9).value = i + 0.8
        row += 1
    #10
    row = 0
    col = 0
    for i in range(0, 8190, 1):
        ws.cell(row=row + 1, column=col + 10).value = i + 0.9
        row += 1
    ex.save("newdata.xlsx")
    print('save success!')


main()
